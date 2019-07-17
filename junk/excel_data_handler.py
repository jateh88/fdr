import openpyxl
import pathlib


def is_integer(string):
    integers = str(1234567890)
    if string in integers:
        return True
    else:
        return False


def get_first_integer_sequence(value):
    """Search through string, return first set of consecutive numbers as single integer

    :param value:
    :return: integer. `0` if no integer was found
    """

    first_integer_set_found = False
    integer_string = ''
    output = -1  # default
    try:
        for char in value:
            if is_integer(char):
                # print(char)
                first_integer_set_found = True
                integer_string += char
            elif first_integer_set_found:
                break
            else:
                pass
        if integer_string != '':
            output = int(integer_string)
    finally:
        return output


def convert_to_string_with_leading_zeroes(value, min_length=0):
    """
    `84, 3` --> `084`
    `321, 1` --> `321`
    `hello, 3` --> `hello`
    :param value:
    :param min_length:
    :return:
    """
    output = '0' * min_length  # default output
    try:
        input_value = str(value)
        length = len(input_value)
        # print(f'length = {length}')
        missing_length = min_length - length
        if missing_length > 0:
            leading_zeroes = '0' * missing_length
            # print(f'leading zeroes = {leading_zeroes}')
            output = leading_zeroes + input_value
        else:
            output = input_value
    finally:
        return output


def get_dicts_from_excel(path):
    """Return a dictionary. Keys = worksheet names; Values = worksheet contents
    Those Values are themselves lists of dictionaries. Each list item is a worksheet row.
    Each worksheet row is a dictionary. Keys = column names; Values = cell values

    :param path: Should be a pathlib.Path() object

    """

    wb = openpyxl.load_workbook(path)
    workbook_contents = dict()

    for worksheet_name in wb.sheetnames:
        ws = wb[worksheet_name]
        worksheet_contents = []
        # Convert each row to a dictionary, using the column headers as keys
        for row in range(2, ws.max_row + 1):
            row_contents = dict()
            for col in range(1, ws.max_column + 1):
                key = ws.cell(1, col).value
                value = ws.cell(row, col).value
                row_contents[key] = value
            worksheet_contents.append(row_contents)
        workbook_contents[worksheet_name] = worksheet_contents

    return workbook_contents

def get_boolean

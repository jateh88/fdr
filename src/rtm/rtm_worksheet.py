import openpyxl
import click
from rtm.fields import Field, WorksheetColumn, field_classes as fc
from rtm.exceptions import RTMValidatorFileError
from typing import List


class RTMWorksheet:

    # Initialize each field with its data
    def __init__(self, path):
        worksheet_columns = self._get_worksheet_columns(
            path=path, worksheet_name="Procedure Based Requirements"
        )
        self.fields = self._initialize_fields(fc, worksheet_columns)

    @staticmethod
    def _get_worksheet_columns(path, worksheet_name):
        """Return list of WorksheetColumn objects"""

        # --- Get Workbook ----------------------------------------------------
        wb = openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)

        # --- Get Worksheet ---------------------------------------------------
        ws = None
        for sheetname in wb.sheetnames:
            if sheetname.lower() == worksheet_name.lower():
                ws = wb[sheetname]
        if ws is None:
            raise RTMValidatorFileError(
                f"\nError: The RTM workbook does not contain a '{worksheet_name}' worksheet"
            )

        # --- Convert Worksheet to WorksheetColumn objects --------------------
        ws_data = []
        start_column_num = 1
        for col in range(start_column_num, ws.max_column + 1):
            column_header = ws.cell(1, col).value
            column_body = tuple(
                ws.cell(row, col).value for row in range(2, ws.max_row + 1)
            )
            ws_column = WorksheetColumn(header=column_header, body=column_body)
            ws_data.append(ws_column)

        return ws_data

    @staticmethod
    def _initialize_fields(field_classes, worksheet_columns) -> List[Field]:
        """Get list of field objects that each contain their portion of the worksheet_columns"""
        fields = []
        with click.progressbar(field_classes) as bar:
            #     fc.append(field(worksheet_columns))
            # return
            for field in bar:
                fields.append(field(worksheet_columns))
            # fields = [field(worksheet_columns) for field in bar]
        return fields

    def validate(self):

        # --- Check Field Sorting ---------------------------------------------
        index_current = -1
        for field in self.fields:
            index_current = field.validate_position(index_current)

        # --- Validate Fields and Print Results -------------------------------
        for field in self.fields:
            field.validate()


if __name__ == "__main__":
    pass

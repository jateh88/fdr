# This is a list of classes, each containing the logic for a worksheet field (or grouping of fields):
import openpyxl
import click
# from .fields import field_class_seq
from fdr.fields import field_class_seq as field_classes
from fdr.exceptions import RTMValidatorError, RTMValidatorFileError
from collections import namedtuple

WorksheetColumn = namedtuple("WorksheetColumn", "header values")


class FdrWorksheet:

    # Initialize each field with its data
    def __init__(self, path):
        worksheet = self._get_worksheet(path)
        self._fields = self._initialize_fields(field_classes, worksheet)

    @staticmethod
    def _get_worksheet(path):
        """Return list of WorksheetColumn objects"""

        # --- Get Workbook ----------------------------------------------------
        wb = openpyxl.load_workbook(
            filename=str(path),
            # read_only=True,
            # data_only=True,
        )

        # --- Check for Worksheet ---------------------------------------------
        worksheet_name = "Procedure Based Requirements"
        worksheet_names = [name.lower() for name in wb.sheetnames]
        try:
            worksheet_index = worksheet_names.index(worksheet_name.lower())
        except ValueError:
            raise RTMValidatorFileError(
                f"\nError: The RTM workbook does not contain a '{worksheet_name}' worksheet")

        # --- Convert Worksheet to WorksheetColumn objects --------------------
        ws = wb[worksheet_index]
        ws_data = []
        start_column_num = 1
        for col in range(start_column_num, ws.max_column + 1):
            column_header = ws.cell(1, col).value

            column_body = tuple(
                ws.cell(row, col).value
                for row
                in range(2, ws.max_row + 1))

            # column_body = []
            # for row in range(2, ws.max_row + 1):
            #     column_body.append(ws.cell(row, col).value)

            ws_column = WorksheetColumn(header=column_header, values=column_body)
            ws_data.append(ws_column)

        return ws_data

    @staticmethod
    def _initialize_fields(field_classes, worksheet):
        """Get list of field objects that each contain their portion of the worksheet"""
        return [Field(worksheet) for Field in field_classes]

    def validate(self):
        """Validate each field (e.g. 'ID', 'Devices')"""
        if self._import_successful:
            for field in self._fields:
                field.validate()
        else:
            click.echo("Import was unsuccessful, so there wasn't anything to validate.")

    # --- PROPERTIES ----------------------------------------------------------

    @property
    def fields(self):
        return self._fields


if __name__ == "__main__":
    pass

"""

This file serves as a stand in for server based file reading. get_dicts_from_excel is taken from matchmaker2 and
modified slightly to suit the needs of this file.

"""

#use pip to get all modules.
import random
import logging
import openpyxl
from pathlib import Path
from bin.excel_data_handling.data_fields import ApplicationDataFields


# write a function that gets path to excel file
#def get_path()


def get_dicts_from_excel(path):
    """Get excel applicant. Log status.
    :param path: pathlib.Path() object

    """
    # Return a dictionary. Keys = worksheet names; Values = worksheet contents
    #         Those Values are themselves lists of dictionaries. Each list item is a worksheet row.
    #         Each worksheet row is a dictionary. Keys = column names; Values = cell values

    logging.info(f'Looking for excel applicant in `{path}` ...')
    try:
        wb = openpyxl.load_workbook(path)
        logging.info('Found!\n')
    except:
        logging.error('Not found!\n')
        return False

    workbook_contents = dict()
    for worksheet_name in wb.sheetnames:
        ws = wb[worksheet_name]
        worksheet_contents = []
        # Convert each row to a dictionary, using the column headers as keys
        for row in range(2, ws.max_row + 1):
            row_contents = dict()
            for col in range(1, ws.max_column + 1):
                key = ws.cell(1, col).value
                value = ws.cell(row, col).value
                logging.debug(value)
                row_contents[key] = value
            worksheet_contents.append(row_contents)
        workbook_contents[worksheet_name] = worksheet_contents

    return workbook_contents

import openpyxl
from pathlib import Path


def test_integers_from_excel():
    path = Path(__file__).parent / "test.xlsx"
    wb = openpyxl.load_workbook(filename=path, read_only=True)
    ws = wb["integer"]
    start_row = 1
    end_row = ws.max_row
    for row in range(start_row, end_row + 1):
        assert_type = ws.cell(row, 1).value
        value = ws.cell(row, 2).value
        try:
            value = int(value)
        except (TypeError, ValueError):
            # Value stays as-is
            pass
        if assert_type == "pass":
            assert isinstance(value, int)
        elif assert_type == "fail":
            assert not isinstance(value, int)
        else:
            raise ValueError("Assert Type must be 'pass' or 'fail'.")


def test_string_comparison_approach_from_excel():
    path = Path(__file__).parent / "test.xlsx"
    wb = openpyxl.load_workbook(filename=path, read_only=True)
    ws = wb["strings"]
    start_row = 1
    end_row = ws.max_row
    for row in range(start_row, end_row + 1):
        assert_type = ws.cell(row, 1).value
        cell_value = ws.cell(row, 2).value
        if assert_type == "pass":
            assert cell_value.replace(" ", "").lower().find("w/c") != -1
        elif assert_type == "fail":
            assert not cell_value.replace(" ", "").lower().find("w/c") != -1
# Lessons learned: string comparison is effective as long as cell has some contents. If cell is empty (is None) it will
# trigger an AttributeError if this type of string comparison is conducted because the object has no attribute replace,
# lower or find.

# TODO Jonathan note to Eric: I found out this weekend that you can read the values from excel formulas.
#   I forget the key off the top of my head, but it's something like load_workbook(<path>, keep_values=True)